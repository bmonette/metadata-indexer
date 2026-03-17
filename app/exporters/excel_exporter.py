from dataclasses import asdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import FileRecord
from reporting import generate_summary


ALWAYS_INCLUDE_COLUMNS = [
    "file_name",
    "extension",
    "file_type",
    "full_path",
    "parent_folder",
    "size_bytes",
    "size_kb",
    "created_time",
    "modified_time",
    "scan_status",
    "error_message",
]


def get_used_columns(record_dicts: list[dict]) -> list[str]:
    """
    Return a list of columns to export.
    Core columns are always included.
    Metadata columns are only included if at least one record
    contains a non-empty value for that column.
    """
    if not record_dicts:
        return ALWAYS_INCLUDE_COLUMNS.copy()

    all_columns = list(record_dicts[0].keys())
    used_columns = []

    for column in all_columns:
        if column in ALWAYS_INCLUDE_COLUMNS:
            used_columns.append(column)
            continue

        has_value = any(
            record.get(column) not in (None, "", [])
            for record in record_dicts
        )

        if has_value:
            used_columns.append(column)

    return used_columns


def auto_adjust_column_width(worksheet) -> None:
    """
    Adjust column widths based on max content length.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(records: list[FileRecord], output_path: Path) -> None:
    if not records:
        raise ValueError("No records to export.")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    record_dicts = [asdict(record) for record in records]
    headers = get_used_columns(record_dicts)

    workbook = Workbook()

    # =========================
    # Sheet 1: Metadata Index
    # =========================
    sheet = workbook.active
    sheet.title = "Metadata Index"

    sheet.append(headers)

    for record_dict in record_dicts:
        row = [record_dict.get(header) for header in headers]
        sheet.append(row)

    sheet.freeze_panes = "A2"
    auto_adjust_column_width(sheet)

    # =========================
    # Sheet 2: Summary
    # =========================
    summary = generate_summary(records)

    summary_sheet = workbook.create_sheet(title="Summary")

    summary_sheet.append(["Metric", "Value"])

    summary_sheet.append(["Total Files", summary["total_files"]])
    summary_sheet.append(["Successful Files", summary["success_count"]])
    summary_sheet.append(["Errored Files", summary["error_count"]])
    summary_sheet.append(["Total Size (KB)", summary["total_size_kb"]])

    summary_sheet.append([])

    summary_sheet.append(["File Type", "Count"])

    for file_type, count in summary["file_types"].items():
        summary_sheet.append([file_type, count])

    auto_adjust_column_width(summary_sheet)

    workbook.save(output_path)
